"""
RAG 问答评估报告生成器
- 对比模型回答 vs 官方标准答案
- 支持关键词评分 + 语义相似度（本地 SentenceTransformer）
- 输出：CSV / HTML / 带颜色 Excel
- 复用 retrievers.py 的 find_local_embedder 逻辑

输出目录: experiment/results/report/
"""

import sys
import json
import re
from pathlib import Path
from datetime import datetime
import pandas as pd

# ==============================
# 🔍 复用 retrievers.py 的 find_local_embedder 逻辑
# ==============================
def find_local_embedder(model_name: str, cache_registry_path: str):
    """
    在 cache_registry.json 列出的每个路径中查找模型：
      - 先检查路径本身是否有 model_id.txt（已按你要求修改！）
      - 再检查其子目录
    """
    with open(cache_registry_path, "r", encoding="utf-8") as f:
        cache_dirs = json.load(f)

    for base_dir in cache_dirs:
        base_path = Path(base_dir)
        if not base_path.exists():
            continue

        # 🔹 第一步：检查 base_dir 本身
        model_id_file = base_path / "model_id.txt"
        if model_id_file.exists():
            with open(model_id_file, "r", encoding="utf-8") as f2:
                name_in_file = f2.read().strip()
            if name_in_file == model_name:
                print(f"✅ Found local embedder: {model_name} at {base_path}")
                return str(base_path)

        # 🔹 第二步：再检查子目录（兼容其他结构）
        if base_path.is_dir():
            for candidate in base_path.iterdir():
                if candidate.is_dir():
                    model_id_file = candidate / "model_id.txt"
                    if model_id_file.exists():
                        with open(model_id_file, "r", encoding="utf-8") as f2:
                            name_in_file = f2.read().strip()
                        if name_in_file == model_name:
                            print(f"✅ Found local embedder: {model_name} at {candidate}")
                            return str(candidate)

    raise FileNotFoundError(
        f"Embedding model '{model_name}' not found in any directory listed in {cache_registry_path}. "
        f"Please check model_id.txt content and paths."
    )

# ==============================
# 🧠 加载 SentenceTransformer（使用上述函数）
# ==============================
SCRIPT_DIR = Path(__file__).parent.resolve()
SIMILARITY_AVAILABLE = False
sim_model = None

try:
    from sentence_transformers import SentenceTransformer
    SENTENCE_TRANSFORMERS_INSTALLED = True
except ImportError:
    SENTENCE_TRANSFORMERS_INSTALLED = False

if SENTENCE_TRANSFORMERS_INSTALLED:
    TARGET_MODEL_ID = "sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2"
    BENCHMARK_DIR = SCRIPT_DIR.parent / "benchmark"
    REGISTRY_FILE = BENCHMARK_DIR / "cache_registry.json"

    if not REGISTRY_FILE.exists():
        print(f"❌ cache_registry.json 不存在: {REGISTRY_FILE}")
    else:
        try:
            # ✅ 直接调用复用的函数
            local_model_path = find_local_embedder(TARGET_MODEL_ID, str(REGISTRY_FILE))
            sim_model = SentenceTransformer(
                local_model_path,
                trust_remote_code=False,
                local_files_only=True
            )
            SIMILARITY_AVAILABLE = True
            print(f"✅ 成功加载 SentenceTransformer 模型用于相似度计算")
        except Exception as e:
            print(f"❌ 加载嵌入模型失败: {e}")
else:
    print("ℹ️ 未安装 sentence-transformers，跳过语义相似度计算。")

if not SIMILARITY_AVAILABLE:
    print("📌 相似度功能未启用：后续报告中 similarity_score 将留空。")


# ==============================
# 📊 辅助函数（保持不变）
# ==============================

def compute_keyword_score(generated: str, keywords: list) -> tuple[float, list, list]:
    if not keywords:
        return 1.0, [], []
    text_lower = generated.lower()
    matched = [kw for kw in keywords if kw in text_lower]
    missed = [kw for kw in keywords if kw not in text_lower]
    score = len(matched) / len(keywords)
    return score, matched, missed


def compute_similarity_score(generated: str, reference: str) -> float | None:
    if not SIMILARITY_AVAILABLE or sim_model is None:
        return None
    try:
        from sentence_transformers.util import cos_sim
        embeddings = sim_model.encode([generated, reference], convert_to_tensor=True)
        score = cos_sim(embeddings[0], embeddings[1]).item()
        return max(0.0, min(1.0, score))
    except Exception as e:
        print(f"⚠️ 相似度计算出错: {e}")
        return None


def highlight_keywords(text: str, keywords: list) -> str:
    if not keywords:
        return text.replace("\n", "<br>")
    text_html = (
        text.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace("\n", "<br>")
    )
    for kw in keywords:
        pattern = re.compile(re.escape(kw), re.IGNORECASE)
        text_html = pattern.sub(
            f'<span style="background-color: #d4edda; color: #155724; font-weight: bold;">{kw.upper()}</span>',
            text_html
        )
    return text_html


# ==============================
# 📤 主函数（保持不变）
# ==============================

def main():
    # --- 1. 加载最新模型输出 ---
    raw_dir = SCRIPT_DIR / "results" / "raw_outputs"
    run_files = list(raw_dir.glob("run_*.json"))
    if not run_files:
        raise FileNotFoundError(f"未找到 run_*.json 文件，请先运行 evaluate.py")
    latest_run = max(run_files, key=lambda f: f.stat().st_mtime)
    
    with open(latest_run, 'r', encoding='utf-8') as f:
        model_results = json.load(f)

    # --- 2. 加载标准答案 ---
    gt_path = SCRIPT_DIR / "ground_truth.json"
    with open(gt_path, 'r', encoding='utf-8') as f:
        gt_data = json.load(f)
    ground_truth = {
        item["question"]: {
            "answer": item["answer"],
            "keywords": [kw.lower() for kw in item.get("expected_chunks_keywords", [])]
        }
        for item in gt_data["questions"]
    }

    # --- 3. 构建报告数据 ---
    csv_data = []
    html_rows = []

    for item in model_results:
        query = item["original_query"]
        gt_info = ground_truth.get(query, {"answer": "[未找到标准答案]", "keywords": []})
        
        generated = item.get("answer", "[无回答]")
        keywords = gt_info["keywords"]

        # 关键词评分
        keyword_score, matched, missed = compute_keyword_score(generated, keywords)
        
        # 相似度评分
        similarity_score = compute_similarity_score(generated, gt_info["answer"])

        # CSV 行
        csv_data.append({
            "original_query": query,
            "processed_query": item.get("processed_query", ""),
            "generated_answer": generated,
            "ground_truth_answer": gt_info["answer"],
            "keyword_score": round(keyword_score, 3),
            "matched_keywords": ", ".join(matched),
            "missed_keywords": ", ".join(missed),
            "similarity_score": round(similarity_score, 3) if similarity_score is not None else "",
            "manual_score": ""  # 人工评分列
        })

        # HTML 高亮
        gen_highlighted = highlight_keywords(generated, keywords)
        gt_highlighted = highlight_keywords(gt_info["answer"], keywords)
        html_rows.append(f"""
        <tr>
            <td style="vertical-align: top; width: 20%;"><strong>{query}</strong></td>
            <td style="vertical-align: top; width: 40%;">
                <div style="background: #f8f9fa; padding: 8px; border-radius: 4px; margin-bottom: 6px;">
                    {gen_highlighted}
                </div>
                <div style="font-size: 0.9em; color: #6c757d;">
                    <strong>关键词得分:</strong> {keyword_score:.1%} |
                    <span style="color: green;">命中: {', '.join(matched) or '—'}</span> |
                    <span style="color: red;">缺失: {', '.join(missed) or '—'}</span>
                    {' | <strong>相似度:</strong> ' + f'{similarity_score:.1%}' if similarity_score is not None else ''}
                </div>
            </td>
            <td style="vertical-align: top; width: 40%;">
                <div style="background: #e9ecef; padding: 8px; border-radius: 4px;">
                    {gt_highlighted}
                </div>
            </td>
        </tr>
        """)

    # --- 4. 保存 CSV ---
    report_dir = SCRIPT_DIR / "results" / "report"
    report_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    df = pd.DataFrame(csv_data)
    csv_file = report_dir / f"rag_report_{timestamp}.csv"
    df.to_csv(csv_file, index=False, encoding='utf-8-sig')

    # --- 5. 保存 HTML ---
    avg_keyword = df['keyword_score'].mean()
    avg_similarity = df['similarity_score'].replace('', pd.NA).dropna().astype(float).mean() if SIMILARITY_AVAILABLE else None

    html_content = f"""
    <!DOCTYPE html>
    <html lang="zh-CN">
    <head>
        <meta charset="UTF-8">
        <title>RAG 问答评估报告</title>
        <style>
            body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; margin: 20px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th, td {{ border: 1px solid #dee2e6; padding: 12px; text-align: left; vertical-align: top; }}
            th {{ background-color: #f1f3f5; }}
            tr:nth-child(even) {{ background-color: #fafafa; }}
            .header {{ text-align: center; margin-bottom: 20px; }}
            .summary {{ background: #e2e3e5; padding: 12px; border-radius: 6px; margin-bottom: 20px; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>🔍 RAG 问答评估报告</h1>
            <p>生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            <p>原始结果: {latest_run.name}</p>
        </div>
        <div class="summary">
            <strong>整体表现:</strong> 
            平均关键词得分: {avg_keyword:.1%}
            {f' | 平均语义相似度: {avg_similarity:.1%}' if avg_similarity is not None else ''}
        </div>
        <table>
            <thead>
                <tr>
                    <th>问题</th>
                    <th>模型回答（🟢 高亮命中关键词）</th>
                    <th>标准答案（参考）</th>
                </tr>
            </thead>
            <tbody>
                {''.join(html_rows)}
            </tbody>
        </table>
    </body>
    </html>
    """
    html_file = report_dir / f"rag_report_{timestamp}.html"
    with open(html_file, 'w', encoding='utf-8') as f:
        f.write(html_content)

    # --- 6. 保存带颜色的 Excel ---
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill

        wb = Workbook()
        ws = wb.active
        headers = [
            "original_query", "processed_query", "generated_answer", "ground_truth_answer",
            "keyword_score", "matched_keywords", "missed_keywords", "similarity_score", "manual_score"
        ]
        ws.append(headers)

        for row_data in csv_data:
            row = [
                row_data["original_query"],
                row_data["processed_query"],
                row_data["generated_answer"],
                row_data["ground_truth_answer"],
                row_data["keyword_score"],
                row_data["matched_keywords"],
                row_data["missed_keywords"],
                row_data["similarity_score"],
                row_data["manual_score"]
            ]
            ws.append(row)

            # 设置颜色（基于 keyword_score）
            last_row = ws.max_row
            score = row_data["keyword_score"]
            if score >= 0.8:
                fill = PatternFill(start_color="E6FCE6", end_color="E6FCE6", fill_type="solid")
            elif score >= 0.5:
                fill = PatternFill(start_color="FFF9C2", end_color="FFF9C2", fill_type="solid")
            else:
                fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            
            for cell in ws[last_row]:
                cell.fill = fill

        excel_file = report_dir / f"rag_report_{timestamp}.xlsx"
        wb.save(excel_file)
        print(f"✅ Excel 报告已保存: {excel_file}")

    except ImportError:
        print("⚠️ 未安装 openpyxl，跳过 Excel 生成。请运行: pip install openpyxl")

    # --- 7. 最终提示 ---
    print(f"✅ RAG 评估报告已生成:")
    print(f"   📊 CSV:     {csv_file}")
    print(f"   🌐 HTML:    {html_file}")
    if 'excel_file' in locals():
        print(f"   📈 Excel:   {excel_file}")
    print(f"   📌 平均关键词得分: {avg_keyword:.1%}")


if __name__ == "__main__":
    main()